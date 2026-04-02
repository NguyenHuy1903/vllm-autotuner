"""
Excel Exporter — Xuất kết quả benchmark ra file .xlsx.

3 sheets:
  1. Summary: 1 row/run với các metrics chính
  2. Configs: cấu hình đầy đủ của từng run
  3. Errors: chỉ failed runs với error logs
"""
from __future__ import annotations

import json
import os
import tempfile
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Màu sắc theo status ────────────────────────────────────────────────────────
COLORS = {
    "success": "C6EFCE",    # Xanh lá nhạt
    "failed": "FFC7CE",     # Đỏ nhạt
    "timeout": "FFEB9C",    # Vàng nhạt
    "pending": "DDDDDD",    # Xám
    "running": "BDD7EE",    # Xanh dương nhạt
    "header": "4472C4",     # Xanh dương đậm (header)
    "header_text": "FFFFFF", # Trắng (text header)
}


def _header_style(cell, color: str = "header") -> None:
    """Áp dụng style cho header cell."""
    cell.font = Font(bold=True, color=COLORS["header_text"], size=10)
    cell.fill = PatternFill(fill_type="solid", fgColor=COLORS[color])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _data_style(cell, status: str = "") -> None:
    """Áp dụng style cho data cell theo status."""
    if status in COLORS:
        cell.fill = PatternFill(fill_type="solid", fgColor=COLORS[status])
    cell.alignment = Alignment(vertical="center", wrap_text=False)


def _auto_width(ws, max_width: int = 40) -> None:
    """Tự động điều chỉnh độ rộng cột."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def export_to_xlsx(
    runs: list[dict],
    model_name: Optional[str] = None,
    output_dir: Optional[str] = None,
) -> str:
    """
    Xuất danh sách benchmark runs ra file Excel.

    Args:
        runs: Danh sách dict từ db.get_all_runs_for_export()
        model_name: Tên model (dùng trong tên file)
        output_dir: Thư mục lưu file, mặc định dùng /tmp

    Returns:
        Đường dẫn file .xlsx đã tạo
    """
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    sum_headers = [
        "Run ID", "Model", "TP Size", "Max Num Seqs", "GPU Mem Util",
        "Status", "Throughput (tok/s)", "TTFT (ms)", "Avg Latency (ms)",
        "P99 Latency (ms)", "Max Concurrent", "Startup (s)",
        "Docker Image", "GPU IDs", "Timestamp",
    ]

    ws_sum.row_dimensions[1].height = 30
    for col_idx, header in enumerate(sum_headers, start=1):
        cell = ws_sum.cell(row=1, column=col_idx, value=header)
        _header_style(cell)

    # Freeze header row
    ws_sum.freeze_panes = "A2"

    for row_idx, run in enumerate(runs, start=2):
        status = run.get("status", "")
        values = [
            run.get("run_id", ""),
            run.get("model_name", ""),
            run.get("tensor_parallel_size", ""),
            run.get("max_num_seqs", ""),
            run.get("gpu_memory_utilization", ""),
            status,
            run.get("throughput_tok_s", ""),
            run.get("ttft_ms", ""),
            run.get("avg_latency_ms", ""),
            run.get("p99_latency_ms", ""),
            run.get("max_concurrent_tested", ""),
            run.get("startup_time_s", ""),
            run.get("docker_image", ""),
            run.get("gpu_ids", ""),
            run.get("created_at", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_sum.cell(row=row_idx, column=col_idx, value=val)
            _data_style(cell, status)

    _auto_width(ws_sum)

    # ── Sheet 2: Configs ──────────────────────────────────────────────────────
    ws_cfg = wb.create_sheet("Configs")

    cfg_headers = [
        "Run ID", "Model", "TP Size", "Max Num Seqs", "GPU Mem Util",
        "Max Model Len", "DType", "Docker Image", "GPU IDs",
        "Extra Args", "Status", "Error Class", "Created At",
    ]

    ws_cfg.row_dimensions[1].height = 30
    for col_idx, header in enumerate(cfg_headers, start=1):
        cell = ws_cfg.cell(row=1, column=col_idx, value=header)
        _header_style(cell)

    ws_cfg.freeze_panes = "A2"

    for row_idx, run in enumerate(runs, start=2):
        config_json = run.get("config_json", "{}")
        try:
            cfg = json.loads(config_json) if config_json else {}
        except Exception:
            cfg = {}

        values = [
            run.get("run_id", ""),
            run.get("model_name", ""),
            run.get("tensor_parallel_size", ""),
            run.get("max_num_seqs", ""),
            run.get("gpu_memory_utilization", ""),
            cfg.get("max_model_len", ""),
            cfg.get("dtype", "auto"),
            run.get("docker_image", ""),
            run.get("gpu_ids", ""),
            json.dumps(cfg.get("extra_args", {})),
            run.get("status", ""),
            run.get("error_class", ""),
            run.get("created_at", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_cfg.cell(row=row_idx, column=col_idx, value=val)
            _data_style(cell, run.get("status", ""))

    _auto_width(ws_cfg)

    # ── Sheet 3: Errors ───────────────────────────────────────────────────────
    ws_err = wb.create_sheet("Errors")

    err_headers = [
        "Run ID", "Model", "TP", "Seqs", "GPU Util",
        "Error Class", "Error Log (truncated)", "Timestamp",
    ]

    ws_err.row_dimensions[1].height = 30
    for col_idx, header in enumerate(err_headers, start=1):
        cell = ws_err.cell(row=1, column=col_idx, value=header)
        _header_style(cell)

    ws_err.freeze_panes = "A2"

    failed_runs = [r for r in runs if r.get("status") not in ("success", "pending", "running")]
    for row_idx, run in enumerate(failed_runs, start=2):
        error_log = run.get("error_log", "") or ""
        # Giới hạn 2000 ký tự trong Excel cell
        error_log_trunc = error_log[:2000] if error_log else ""

        values = [
            run.get("run_id", ""),
            run.get("model_name", ""),
            run.get("tensor_parallel_size", ""),
            run.get("max_num_seqs", ""),
            run.get("gpu_memory_utilization", ""),
            run.get("error_class", ""),
            error_log_trunc,
            run.get("created_at", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_err.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 7:  # Error log column
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                _data_style(cell, "failed")

    # Error log column: cho rộng hơn và wrap text
    ws_err.column_dimensions["G"].width = 60
    ws_err.row_dimensions[1].height = 30
    for row_idx in range(2, len(failed_runs) + 2):
        ws_err.row_dimensions[row_idx].height = 60

    _auto_width(ws_err, max_width=30)
    ws_err.column_dimensions["G"].width = 60  # Ghi đè lại cột error log

    # ── Lưu file ──────────────────────────────────────────────────────────────
    if output_dir is None:
        output_dir = tempfile.gettempdir()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    model_tag = model_name.replace("/", "_")[:30] if model_name else "all"
    filename = f"vllm_benchmark_{model_tag}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb.save(filepath)
    return filepath
